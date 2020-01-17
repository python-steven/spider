from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from django.core import serializers
from app import restful,mail
from datetime import datetime,timedelta
from django.conf import settings
import random
import string
import os
import time
import psycopg2
from openpyxl import load_workbook,Workbook
import json


#define this normal
UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# Create your views here.

#login in index
class IndexView(View):
    @csrf_exempt
    def get(self,request):
        try:
            id = request.session['user_Id']
            user = User.objects.get(Id=id)
            if user.IsActivated is False:
                Update_User_IsActivated(id)
                request.session.set_expiry(0)
                num = BudgetCodeForm.objects.filter(Status="Process",PicId=id).count()
                return render(request, "./index/main.html", {'user': user,'num':num})
            else:
                request.session.set_expiry(0)
                num = BudgetCodeForm.objects.filter(Status="Process",PicId=id).count()
                return render(request, "./index/main.html",{'user':user,'num':num})
        except:
            return HttpResponseRedirect("/login/")

    @csrf_exempt
    def post(self,request):
        pass

#修改激活状态工具
def update_User_IsActivated(id):
    User.objects.filter(Id=id).update(IsActivated=True)


#用户管理页面的数据的获取和增加用户：
class UserData(View):


    def get(self,request):
        sql1 = 'SELECT "User"."Id","EmployeeId","Name","Department","Email","Role" FROM "User" INNER JOIN ' \
               '"Department" ON "User"."DepartmentId" = "Department"."Id" WHERE "User"."IsActivated"=True'
        cur = connection.cursor()
        cur.execute(sql1)
        data = cur.fetchall()
        return restful.ok(data=data)

    #增加用户的功能
    @csrf_exempt
    def post(self,request):
            password = genPassword()
            employee_id =request.POST['userid']
            name =request.POST['username']
            email = request.POST['mail']
            role = request.POST['role']
            department =request.POST['department']
            updated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            try:
                user =User.objects.get(Name=name)
                if user:
                    return restful.params_error(message='user name had used')
            except:
                try:
                    user = User.objects.get(EmployeeId=employee_id)
                    if user:
                        return restful.params_error(message="user employeeId had used")
                except:
                    try:
                        user =User.objects.get(Email=email)
                        if user:
                            return restful.params_error(message="user email had used")
                    except:
                        try:
                            department_ob = Department.objects.get(Department=department)
                            department_id = department_ob.Id
                            if department_id:
                                User.objects.create(EmployeeId=employee_id, Name=name, Password=password, Email=email,
                                                    Role=role,CreatedTime=created_time, UpdatedTime=updated_time,
                                                    DepartmentId=department_id)
                                subject = "Notification: AEMSLite account information"
                                content = """
<pre>
Dear """ + name + """,
    Your AEMSLite account and password are as follows:
    account number:""" + name + """
    password:""" + password + """
    Please click the link below to activate the account:
    <a href="http://10.41.95.106:90/login">Login AEMSLite</a>
    
    
    =====================================================================================================
    THIS EMAIL WAS SENT BY SMCS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""

                                mail.sendmail([email,], content, subject)
                                return restful.ok(message='User created success')
                        except:
                            User.objects.create(EmployeeId=employee_id, Name=name, Password=password, Email=email,
                                                Role=role,CreatedTime=created_time, UpdatedTime=updated_time)
                            subject = "Notification: AEMSLite account information"
                            content = """
<pre>
Dear """+name+""",
    Your AEMSLite account and password are as follows:
    account number:"""+name+"""
    password:"""+password+"""
                                
    Please click the link below to activate the account:
    <a href="http://10.41.95.106:90/login">Login AEMSLite</a>
                                
    =====================================================================================================
                                
    THIS EMAIL WAS SENT BY SMCS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
 """
                            mail.sendmail([email,], content,subject)
                            return restful.ok(message='User add success')

#随机生成密码的函数
def genPassword(length=8, chars=string.digits + string.ascii_letters):
    return ''.join(random.sample(chars * 10, 8))

#修改用户的相关信息的函数
@csrf_exempt
def modify_user(request):
    if request.method == "POST":
        id = int(request.POST['id'])
        employee_id = request.POST['employee_id']
        username = request.POST['username']
        department = request.POST['department']
        email = request.POST['email']
        role = request.POST['role']
        try:
            original__employee_id = User.objects.exclude(Id=id).get(EmployeeId=employee_id)
            if original__employee_id:
                return restful.params_error(message='employee had used')
        except:
            try:
                original_name =User.objects.exclude(Id=id).get(Name=username)
                if original_name:
                    return restful.params_error(message='user name had used')
            except:
                try:
                    original_email = User.objects.exclude(Id=id).get(Email=email)
                    if original_email:
                        return restful.params_error(message='email had used')
                except:
                    try:
                        modify_department = Department.objects.get(Department=department)
                        department_id = modify_department.Id
                        if modify_department:
                            User.objects.filter(Id=id).update(EmployeeId=employee_id, Name=username,Email=email,
                                                              Role=role,UpdatedTime=UpdatedTime,
                                                              DepartmentId=department_id)
                            return restful.ok(message="user modify success")
                    except:
                        User.objects.filter(Id=id).update(EmployeeId=employee_id, Email=email,Name=username,
                                                          Role=role,UpdatedTime=UpdatedTime)
                        return restful.ok(message='User modify success')

                    return restful.params_error(message='email had used')
                return restful.params_error(message='user name had used')
            return restful.params_error(message='employee had used')

#删除用户的函数
@csrf_exempt
def del_user(request):
    if request.method == "POST":
        name = request.POST['name']
        try:
            user = User.objects.get(Name=name)
            if user:
                user.IsActivated = False
                user.save()
                return restful.ok(message='delete success')
        except:
            return restful.params_error(message='delete user fail')


#增加客户和获取客户的信息
class CustomerInfo(View):

    def get(self,request):
        customerinfo = Customer.objects.exclude(IsActivated='False')
        customerinfo = customerinfo.values()
        customerinfo = list(customerinfo)
        return restful.ok(data=customerinfo)

    @csrf_exempt
    def post(self,request):
        customer = request.POST['customer_val']
        updatedtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            customer_ob = Customer.objects.get(Customer=customer)
            if customer_ob.Id:
                return restful.params_error(message="Customer had Exist")
        except:
            Customer.objects.create(Customer=customer,UpdatedTime=updatedtime)
            return restful.ok(message='Customer add success')


#修改客户数据
def modify_customer(request):
    if request.method == "POST":
        customer = request.POST['customer_name']
        customer_id = request.POST['customer_id']
        try:
            customer_ob = Customer.objects.exclude(Id=customer_id).get(Customer=customer)
            if customer_ob:
                return restful.params_error(message="Customer had used")
        except:
            Customer.objects.filter(Id=customer_id).update(Customer=customer,UpdatedTime=UpdatedTime)
            return restful.ok(message="customer had modify")

#删除客户数据
def del_customer(request):
    if request.method == "POST":
        customer = request.POST['del_nm']
        try:
            cus = Customer.objects.get(Customer=customer)
            if cus:
                cus.IsActivated = False
                cus.save()
                return restful.ok(message='delete success')
        except:
            return restful.params_error(message='delete Customer fail')


#增加部门和获取部门数据
class DepartmentInfo(View):

    def get(self,request):
        department_info = Department.objects.exclude(IsActivated='False')
        department_info = department_info.values()
        department_info = list(department_info)
        return restful.ok(data=department_info)

    @csrf_exempt
    def post(self,request):
        department = request.POST['department']
        updatedtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            part = Department.objects.get(Department=department)
            if part.Id:
                return restful.params_error(message="Department had Exist")
        except:
            Department.objects.create(Department=department,UpdatedTime=updatedtime)
            return restful.ok(message='Department create success')


#修改部门数据
def modify_department(request):
    if request.method == "POST":
        depart = request.POST['modifyPartName']
        depart_id = request.POST['modifyPartId']
        try:
            depart_ob = Department.objects.exclude(Id=depart_id).get(Department=depart)
            if depart_ob:
                return restful.params_error(message="Customer had used")
        except:
            Department.objects.filter(Id=depart_id).update(Department=depart,UpdatedTime=UpdatedTime)
            return restful.ok(message="Department had modify")

#删除部门数据
def delete_department(request):
    if request.method == "POST":
        depart_name = request.POST['delPart']
        try:
            depart = Department.objects.get(Department=depart_name)
            if depart:
                depart.IsActivated = False
                depart.save()
                return restful.ok(message='delete success')
        except:
            return restful.params_error(message='delete Department fail')


#get customer and department info
def Budget_info_get(request):
    if request.method == "GET":
        data_dict={}
        cus_info = Customer.objects.exclude(IsActivated='False')
        cus_info = list(cus_info.values())
        depart_info = Department.objects.exclude(IsActivated='False')
        depart_info = list(depart_info.values())
        data_dict['customer'] =cus_info
        data_dict['department'] =depart_info
        return restful.ok(data=data_dict)

#check user and principal
def Budget_check_user(request):
    if request.method == "POST":
        check_user = request.POST['user_approve']
        try:
            user_c = User.objects.get(Name=check_user)
            if user_c:
                return restful.ok()
        except:
            return restful.params_error(message='the Signer not exist need admin check')

def Budget_check_principal(request):
    if request.method == "POST":
        principal = request.POST['principal']
        try:
            user_c = User.objects.get(Name=principal)
            if user_c:
                return restful.ok()
        except:
            try:
                user_num = User.objects.get(EmployeeId=principal)
                if user_num:
                    return restful.ok()
            except:
                return restful.params_error(message='the PIC user not exist need admin check')




#Budgetcode application form
class BudgetCodeApply(View):

    def get(self,request):
        try:
            id = request.session['user_Id']
            budgetcode_info = BudgetCodeForm.objects.filter(PicId=id).values("Id","BillingType","Department","ApplyDate"
                                                                  ,"Pic","ProductName","Signer","Status","BudgetCode")

            budgetcode_info = list(budgetcode_info)
            return restful.ok(data=budgetcode_info)
        except:
            return restful.params_error(message="")

    @csrf_exempt
    def post(self,request):
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal')
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        bud_request_type = request.POST.get('bud_request_type')
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        # bud_total_price = request.POST['bud_total_price']
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_project_code = request.POST.get('bud_project_code')
        #
        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # upload_tool = request.POST.get('upload_tool')
        own_id = request.session.get('user_Id')
        # name_current = own_id.Name

        # url_num = genurlnum()
        time_num = int(time.time())
        time_num = str(time_num)
        file = request.FILES.get('upload_file')
        file_name = file.name
        file_sp_name =file_name.split('.')[0]
        file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]

        file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        try:
            depart = Department.objects.get(Department=bud_depart)
            depart_id = depart.Id
            try:
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                try:
                    cus = Customer.objects.get(Customer=bud_customer)
                    cus_id = cus.Id
                    try:
                        pic_user = User.objects.get(Name=bud_principal)
                        pic_user_id = pic_user.Id
                        if pic_user.IsActivated is True:
                            BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                                  , Attachment=file_sp_name, ApplyDate=created_time
                                                  , ExternalNumberEffectiveDate=bud_time
                                                  , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                  , PicId=pic_user_id,Pic=bud_principal, ProductName=bud_machine_name
                                                  , Model=bud_machine_type,PurchaseType=bud_request_type
                                                  , UnitPrice=bud_price, Quantity=bud_qty,Unit=bud_qty_type
                                                  , Currency=bud_money_type, CustomerId=cus_id
                                                  , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                                  , ProjectCode=bud_project_code, ApplyReason=bud_reason,SignerId=user_id
                                                  , Signer=bud_user, Status='Process', CreatedTime=created_time
                                                  , UpdatedTime=UpdatedTime, OwnerId=own_id,AttachmentPath=file_url
                                                  )
#邮件发送创建的表单给签核的人去签核表单信息
                            subject = "Notification: AEMSLite Budget_code signing form information"
                            email_1 = user.Email
                            email_2 = pic_user.Email
                            content = """
<pre>
Dear """ + bud_user + """,
    You have a Budget_Code Form application need you sign and please seen the below link address:
    Budget_code apply department:""" + bud_depart + """    
    Budget_code apply charger:""" + bud_principal + """
    Budget_code apply product name:""" + bud_machine_name + """
    Please click the link below to signed the apply:
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>


    =====================================================================================================
    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                            mail.sendmail([email_1,email_2], content, subject)
                            return restful.ok(message='BudgetCodeForm create success and mail had send out')
                        else:
                            return restful.ok(message="PIC user had not activate")
                    except:
                        try:
                            pic_user_2 = User.objects.get(EmployeeId=bud_principal)
                            pic_user_2_id = pic_user_2.Id
                            if pic_user_2.IsActivated is True:
                                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart,Remark=bud_req
                                                              , Attachment=file_sp_name, ApplyDate=created_time
                                                              , ExternalNumberEffectiveDate=bud_time
                                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                              , PicId=pic_user_2_id, Pic=bud_principal
                                                              , ProductName=bud_machine_name
                                                              , Model=bud_machine_type, PurchaseType=bud_request_type
                                                              , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                                              , Currency=bud_money_type, CustomerId=cus_id
                                                              , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                                              , ProjectCode=bud_project_code, ApplyReason=bud_reason
                                                              , SignerId=user_id, Signer=bud_user, Status='Process'
                                                              , CreatedTime=created_time
                                                              , UpdatedTime=UpdatedTime, OwnerId=own_id
                                                              , AttachmentPath=file_url
                                                              )
                                # 邮件发送修改内容
                                subject = "Notification: AEMSLite Budget_code signing form information"
                                email_1 = user.Email
                                email_2 = pic_user_2.Email
                                content = """
<pre>
Dear """ + bud_user + """,
You have a Budget_Code Form application need you sign and please seen the below link address:
Budget_code apply department:""" + bud_depart + """                                    
Budget_code apply charger:""" + bud_principal + """
Budget_code apply product name:""" + bud_machine_name + """
Please click the link below to signed the apply:
<a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>


=====================================================================================================
THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                                ret = mail.sendmail([email_1, email_2], content, subject)
                                if ret:
                                    return restful.ok(message='BudgetCodeForm create success and mail had send out')
                                else:
                                    return restful.params_error(message='mail send fail')
                            else:
                                return restful.ok(message="PIC user had not activate")
                        except:
                            return restful.params_error(message="PIC need username or employee")
                except:
                    return restful.params_error(message="Customer no exist")
            except:
                return restful.params_error(message="user no exist")
        except:
            return restful.params_error(message="department no exist")

#save form to db
def Budget_form_save(request):
    if request.method == "POST":
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal')
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        bud_request_type = request.POST.get('bud_request_type')
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        # bud_total_price = request.POST['bud_total_price']
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_project_code = request.POST.get('bud_project_code')
        #
        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # upload_tool = request.POST.get('upload_tool')
        own_id = request.session.get('user_Id')

        # url_num = genurlnum()
        time_num = int(time.time())
        time_num =str(time_num)

        file = request.FILES.get('upload_file')
        file_name = file.name
        file_sp_name = file_name.split('.')[0]
        file_ven_name = file_sp_name +time_num +'.' + file_name.split('.')[1]

        file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        try:
            depart = Department.objects.get(Department=bud_depart)
            depart_id = depart.Id
            try:
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                try:
                    cus = Customer.objects.get(Customer=bud_customer)
                    cus_id = cus.Id
                    try:
                        pic_user = User.objects.get(Name=bud_principal)
                        pic_user_id = pic_user.Id
                        if pic_user.IsActivated is True:
                            BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                                          , ApplyDate=created_time, Attachment=file_sp_name
                                                          , ExternalNumberEffectiveDate=bud_time
                                                          , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                          , PicId=pic_user_id, Pic=bud_principal
                                                          , ProductName=bud_machine_name,Model=bud_machine_type
                                                          , PurchaseType=bud_request_type, UnitPrice=bud_price
                                                          , Quantity=bud_qty, Unit=bud_qty_type
                                                          , Currency=bud_money_type, CustomerId=cus_id
                                                          , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                                          , ProjectCode=bud_project_code, ApplyReason=bud_reason
                                                          , SignerId=user_id,Signer=bud_user, Status='Draft'
                                                          , CreatedTime=created_time, UpdatedTime=UpdatedTime
                                                          , OwnerId=own_id,AttachmentPath=file_url
                                                          )
                            return restful.ok(message="BudgetCodeForm create success")
                        else:
                            return restful.ok(message="PIC user had not activate")
                    except:
                        try:
                            pic_user_2 = User.objects.get(EmployeeId=bud_principal)
                            pic_user_2_id = pic_user_2.Id
                            if pic_user_2.IsActivated is True:
                                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart
                                                              , Remark=bud_req,ApplyDate=created_time
                                                              , Attachment=file_sp_name
                                                              , ExternalNumberEffectiveDate=bud_time
                                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                              , PicId=pic_user_2_id, Pic=bud_principal
                                                              , ProductName=bud_machine_name
                                                              , Model=bud_machine_type, PurchaseType=bud_request_type
                                                              , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                                              , Currency=bud_money_type, CustomerId=cus_id
                                                              , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                                              , ProjectCode=bud_project_code, ApplyReason=bud_reason
                                                              ,SignerId=user_id, Signer=bud_user, Status='Draft'
                                                              ,CreatedTime=created_time,UpdatedTime=UpdatedTime
                                                              ,OwnerId=own_id,AttachmentPath=file_url
                                                              )
                                return restful.ok(message="BudgetCodeForm create success")
                            else:
                                return restful.ok(message="PIC user had not activate")
                        except:
                            return restful.params_error(message="PIC need username or employee")
                except:
                    return restful.params_error(message="Customer no exist")
            except:
                return restful.params_error(message="user no exist")
        except:
            return restful.params_error(message="department no exist")

#megre order information get and Rendering to html
def Budget_merge_order(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budgetcode_megre_info = BudgetCodeForm.objects.filter(Status='Draft',OwnerId=id).values("Id","Department", "ApplyDate", "Pic",
                                                            "ProductName","Signer", "Status")
            budgetcode_megre_info =list(budgetcode_megre_info)
            return restful.ok(data=budgetcode_megre_info)
        except:
            return restful.params_error(message="")

#megre order post data to signer to sign this form
def merge_form_sub(request):
    if request.method == "POST":
        checked_id_array = request.POST.getlist('ids[]')
        time_id =int(time.time())
        # return restful.params_error(data={'data': checked_id_array})
        try:
            budget_user = BudgetCodeForm.objects.filter(Id__in=checked_id_array).values("Signer").distinct()
            singer_num = budget_user.count()
            if singer_num == 1:
                BudgetCodeForm.objects.filter(Id__in=checked_id_array).update(MergeId=time_id, BillingType=1,Status="Process")
                # for i in checked_id_array:
                # BudgetCodeForm.objects.filter(Id=i).update(MergeId=time_id,BillingType=1,Status="Process")
                # budget_user = BudgetCodeForm.objects.get(Id=i)
                # budget_user = list(budget_user.values())

                # 邮件发送合并表单的要签核的信息给签核人
                subject = "Notification: AEMSLite Budget_code merged form signing form information"
                # budget_code_department = budget_user.Department
                # budget_code_principal = budget_user.Pic
                # budget_code_ProductName = budget_user.ProductName
                # budget_user_name = budget_user.Signer
                budget_user =list(budget_user)
                sign_user = User.objects.get(Name=budget_user[0]['Signer'])
                email_1 = sign_user.Email
                sign_user_name = sign_user.Name
                content = """
<pre>
Dear """ + sign_user_name + """,
    You have a Budget_Code merged Form application need you sign and please seen the below link address:
    
    Please click the link below to signed the apply:
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    
    
    =====================================================================================================
    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                mail.sendmail([email_1,], content, subject)
                return restful.ok(message='BudgetCodeForm merged success')
            return restful.params_error(message="merged form signer different")
        except:
            return restful.params_error(message="merged form fail")


#修改表单信息
def budget_modify_type(request):
    if request.method == "POST":
        modify_id = request.POST['modify_id']
        modify_date = request.POST['modify_date']
        modify_number = request.POST['modify_number']
        try:
            modify_budget_ob =BudgetCodeForm.objects.get(Id=modify_id)
            try:
                mer_id =modify_budget_ob.MergeId
                BudgetCodeForm.objects.filter(MergeId=mer_id).update(ExternalNumberEffectiveDate=modify_date,
                                                                     ExternalNumber=modify_number)
            except:
                BudgetCodeForm.objects.filter(Id=modify_id).update(ExternalNumberEffectiveDate=modify_date, ExternalNumber=modify_number)
                return restful.ok(message="modify form success")
        except:
            return restful.params_error(message="data error")
#修改reject和draft的表单的获取信息
def budget_modify_unique(request):
    if request.method == "POST":
        modify_unique_id = request.POST['modify_unique_id']
        try:
            modify_unique_ob = BudgetCodeForm.objects.filter(Id=modify_unique_id)
            modify_unique_ob = list(modify_unique_ob.values())
            return restful.ok(data=modify_unique_ob)
        except:
            return restful.params_error(message="data not exist!")


#复制表单信息
def budget_copy_type(request):
    if request.method == "POST":
        copy_id = request.POST['copy_id']
        try:
            cop_ob = BudgetCodeForm.objects.get(Id=copy_id)
            BudgetCodeForm.objects.create(DepartmentId=cop_ob.DepartmentId, Department=cop_ob.Department
                                          , Remark=cop_ob.Remark, ApplyDate=cop_ob.ApplyDate
                                          , Attachment=cop_ob.Attachment
                                          , ExternalNumberEffectiveDate=cop_ob.ExternalNumberEffectiveDate
                                          , ExternalNumberType=cop_ob.ExternalNumberType, ExternalNumber=cop_ob.ExternalNumber
                                          , PicId=cop_ob.PicId, Pic=cop_ob.Pic
                                          , ProductName=cop_ob.ProductName
                                          , Model=cop_ob.Model, PurchaseType=cop_ob.PurchaseType
                                          , UnitPrice=cop_ob.UnitPrice, Quantity=cop_ob.Quantity, Unit=cop_ob.Unit
                                          , Currency=cop_ob.Currency, CustomerId=cop_ob.CustomerId
                                          , Customer=cop_ob.Customer, TypeOfMachine=cop_ob.TypeOfMachine
                                          , ProjectCode=cop_ob.ProjectCode, ApplyReason=cop_ob.ApplyReason
                                          , SignerId=cop_ob.SignerId, Signer=cop_ob.Signer, Status='Draft'
                                          , CreatedTime=cop_ob.CreatedTime, UpdatedTime=cop_ob.UpdatedTime
                                          , OwnerId=cop_ob.OwnerId, AttachmentPath=cop_ob.AttachmentPath
                                          )
            return restful.ok(message='copy form success')
        except:
            return restful.params_error(message="data got fail")

#删除(取消)表单信息
def budget_delete_type(request):
    if request.method == "POST":
        del_id = request.POST['del_id']
        try:
            budget_ob = BudgetCodeForm.objects.get(Id=del_id)
            if budget_ob.Status == "Draft" or budget_ob.Status == "Reject" or budget_ob.Status == "Cancel":
                BudgetCodeForm.objects.filter(Id=del_id).delete()
                return restful.ok(message="form delete success")

            if budget_ob.Status == "Process":
                BudgetCodeForm.objects.filter(Id=del_id).update(Status="Cancel")
                cancel_budget = BudgetCodeForm.objects.get(Id=del_id)
                cancel_budget_Name =cancel_budget.Signer
                cancel_budget_principal= cancel_budget.Pic

                mail_user = User.objects.get(Name=cancel_budget_Name)
                mail_principal = User.objects.get(Name=cancel_budget_principal)
# 此表单为cancel的表单提交信息 发送email给相关的签核以及负责人
                email_1 = mail_user.Email
                email_2 = mail_principal.Email
                subject = "Notification: AEMSLite Budget_code Cancel form information"
                content = """
<pre>
Dear """ + cancel_budget_Name + """,
You have a Budget_Code Form had cancel and detail information please seen the below link address:

Please click the link below to signed the apply:
<a href="http://10.41.95.106:90/index">index AEMSLite</a>


=====================================================================================================
THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
 """
                mail.sendmail([email_1, email_2], content, subject)
                return restful.ok(message='BudgetCodeForm cancel success and mail had send out')
                # return restful.ok(message="form delete success")
            # return restful.params_error(message="this form had delete")
        except:
            return restful.params_error(message="data error")




#签核单号的信息的获取函数
def budget_singing_info(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budget_singing_data = BudgetCodeForm.objects.filter(Status='Process',SignerId=id).values("Id","BillingType", "Department"
                                                                                       , "ApplyDate","Pic","ProductName"
                                                                                       , "Signer", "Status","MergeId")
            budget_singing_data = list(budget_singing_data)
            return restful.ok(data=budget_singing_data)
        except:
            return restful.params_error(message="")

#签核表单内容
def merge_signed(request):
    if request.method == "POST":
        bud_id = request.POST['bud_id']
        bud_merged_id = request.POST['bud_merged_id']
        bud_budgetcode = request.POST['budget_cod_text']
        bud_signremarks = request.POST['budget_text']
        sign_id = request.session['user_Id']
        try:
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)
            if user.Name == sign_budget_ob.Pic:
                if bud_merged_id == "null":
                    bud_obj = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_obj.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode, SignRemarks=bud_signremarks,Status='Approve')
                        sign_user = BudgetCodeForm.objects.get(Id=bud_id)
                        sign_pincipal = sign_user.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
    # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Notification: AEMSLite Budget_code Approved form information"
                        content = """
    <pre>
    Dear """ + sign_pincipal + """,
        You have a Budget_Code Form had approved and detail information please seen the below link address:
        
        Please click the link below to see the detail information:
        <a href="http://10.41.95.106:90/index">index AEMSLite</a>
    
    
    =====================================================================================================
    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    </pre>
     """
                        mail.sendmail([email_1,], content, subject)
                        return restful.ok(message='BudgetCodeForm approved and mail had send out')
                        # return restful.ok(message="form signed")
                else:
                    bud_ob_merged = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_ob_merged.Status == "Cancel":
                        return restful.params_error("this form had Cancel")
                    else:
                        BudgetCodeForm.objects.filter(MergeId=bud_merged_id).update(BudgetCode=bud_budgetcode,
                                                                                    SignRemarks=bud_signremarks,Status='Approve')
                        sign_user = BudgetCodeForm.objects.get(Id=bud_id)
                        sign_pincipal = sign_user.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Notification: AEMSLite Budget_code Approved form information"
                        content = """
    <pre>
    Dear """ + sign_pincipal + """,
        You have a Budget_Code Form had approved and detail information please seen the below link address:
    
        Please click the link below to see the detail information:
        <a href="http://10.41.95.106:90/index">index AEMSLite</a>
    
    
        =====================================================================================================
        THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    </pre>
    """
                        mail.sendmail([email_1,], content, subject)
                        return restful.ok(message='BudgetCodeForm approved and mail had send out')
            else:
                return restful.params_error(message="this signer can not signing")
        except:
            return restful.params_error("information data error")
#拒绝签核表单信息内容
def merge_rejected(request):
    if request.method == "POST":
        bud_id = request.POST['bud_id']
        bud_merged_id = request.POST['bud_merged_id']
        bud_budgetcode = request.POST['budget_cod_text']
        bud_signremarks = request.POST['budget_text']
        sign_id = request.session['user_Id']
        try:
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)
            if user.Name == sign_budget_ob.Pic:
                if bud_merged_id == "null":
                    bud_obj = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_obj.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode, SignRemarks=bud_signremarks,Status='Reject')
                        sign_user = BudgetCodeForm.objects.get(Id=bud_id)
                        sign_user_reason = sign_user.Remark
                        sign_pincipal = sign_user.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Notification: AEMSLite Budget_code rejected form information"
                        content = """
<pre>
Dear """ + sign_pincipal + """,
    You have a Budget_Code Form had rejected and detail information please seen the below link address:
    Budget_code apply reject reason:""" + sign_user_reason + """ 
    Please click the link below to see the detail information:
    <a href="http://10.41.95.106:90/index">index AEMSLite</a>


    =====================================================================================================
    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    </pre>
"""
                        mail.sendmail([email_1, ], content, subject)
                        return restful.ok(message='BudgetCodeForm rejected and mail had send out')
                        # return restful.ok(message="form rejected")
                        #这里需要发邮件告知负责人知晓
                else:
                    bud_ob_merged = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_ob_merged.Status == "Cancel":
                        return restful.params_error("this form had Cancel")
                    else:
                        BudgetCodeForm.objects.filter(MergeId=bud_merged_id).update(BudgetCode=bud_budgetcode,
                                                                                    SignRemarks=bud_signremarks,Status='Reject')
                        sign_user = BudgetCodeForm.objects.get(Id=bud_id)
                        sign_user_reason = sign_user.Remark
                        sign_pincipal = sign_user.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Notification: AEMSLite Budget_code rejected form information"
                        content = """
    <pre>
    Dear """ + sign_pincipal + """,
        You have a Budget_Code Form had rejected and detail information please seen the below link address:
        Budget_code apply reject reason:""" + sign_user_reason + """ 
        Please click the link below to see the detail information:
        <a href="http://10.41.95.106:90/index">index AEMSLite</a>
    
    
        =====================================================================================================
        THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    </pre>
    """
                        mail.sendmail([email_1, ], content, subject)
                        return restful.ok(message='BudgetCodeForm rejected and mail had send out')
                        # return restful.ok(message="merged form reject")
            else:
                return restful.params_error(message="this signer can not signing")
        except:
            return restful.params_error("information data error")



#获取预算编码的内容get
def merge_signed_finished(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budget_singed_info = BudgetCodeForm.objects.filter(Status='Approve',SignerId=id)
            budget_singed_info = budget_singed_info.values("Id","BillingType", "Department", "ApplyDate"
                                                           , "Pic","ProductName", "Signer", "Status"
                                                           , "BudgetCode", "MergeId")
            budget_singed_info_2 = BudgetCodeForm.objects.filter(Status='Reject',SignerId=id)
            budget_singed_info_2 = budget_singed_info_2.values("Id", "BillingType", "Department", "ApplyDate"
                                                               , "Pic", "ProductName", "Signer", "Status"
                                                               , "BudgetCode", "MergeId")
            budget_singed_info = list(budget_singed_info)
            budget_singed_info_2 = list(budget_singed_info_2)
            budget_singed_info = budget_singed_info+budget_singed_info_2
            return restful.ok(data=budget_singed_info)
        except:
            return restful.params_error(message="")


#获取预算编码信息并且生成报表的信息
def merge_statement_detail(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budget_statement_detail = BudgetCodeForm.objects.filter(Status='Approve',SignerId=id)
            budget_statement_detail = budget_statement_detail.values("Id","BillingType", "Department", "ApplyDate"
                                                                    ,"Pic","ProductName", "Signer", "Status"
                                                                    ,"BudgetCode", "MergeId")
            # budget_statement_detail_1 = BudgetCodeForm.objects.filter(Status='Reject',SignerId=id)
            # budget_statement_detail_1 = budget_statement_detail_1.values("Id", "BillingType","Department"
            #                                                              , "ApplyDate", "Pic","ProductName"
            #                                                              , "Signer", "Status","BudgetCode"
            #                                                              , "MergeId")
            # budget_statement_detail_1 = list(budget_statement_detail_1)
            budget_statement_detail = list(budget_statement_detail)
            # budget_statement_detail = budget_statement_detail+budget_statement_detail_1
            return restful.ok(data=budget_statement_detail)
        except:
            return restful.params_error(message=" ")


#筛选信息的获取
def statement_query(request):
    if request.method == "POST":
        query_billing_type = request.POST['query_billing_type']
        query_department = request.POST['query_department']
        query_start_date = request.POST['query_start_date']
        query_end_date = request.POST['query_end_date']
        query_pic = request.POST['query_pic']
        query_product_name = request.POST['query_product_name']
        query_signer = request.POST['query_signer']
        query_status = request.POST['query_status']

        try:
            sql ='select "Id","BillingType","Department","ApplyDate","Pic","ProductName","Signer","Status","BudgetCode"' \
                 ',"MergeId" FROM "BudgetCodeForm" WHERE "Status"=\'Approve\'' #查询状态为approve的对象
            if query_billing_type != "":
                sql = sql + 'AND "Pic" = \''+query_billing_type +'\''
            if query_department != "":
                sql = sql+'and "Department" = \''+ query_department +'\''
            if query_pic != "":
                sql = sql + 'AND "Pic" = \''+query_pic +'\''
            if query_product_name != "":
                sql = sql + 'AND "ProductName" = \''+query_product_name +'\''
            if query_signer != "":
                sql = sql + 'AND "Signer" = \''+query_signer +'\''
            if query_status != "":
                sql = sql + 'AND "Status" = \''+query_status +'\''
            if query_start_date != "":
                sql = sql + 'and "ApplyDate" >=\'' + query_start_date + '\''
            if query_end_date != "":
                sql = sql + 'and "ApplyDate" <=\'' + query_end_date + '\''
            # data_1=[]
            cur = connection.cursor()
            cur.execute(sql)
            data = cur.fetchall()
            # data_2 = list(data)
            # keys = ["Id","BillingType","Department","ApplyDate","Pic","ProductName","Signer","Status","BudgetCode","MergeId"]
            # for a in len(data):
            #     data_1[a] = dict(zip(keys,data[a]))
            return restful.ok(data=data)
            # return restful.params_error(data={'data':data})
        except:
            return restful.params_error(message="query data null")

#详细信息的展示和加载
def budget_code_detail(request):
    if request.method == "POST":
        detail_id = request.POST['detail_id']
        detail_merged_id = request.POST['detail_merged_id']
        detail_merged_id =str(detail_merged_id)
        try:
            try:
                detail_data = BudgetCodeForm.objects.filter(MergeId=detail_merged_id).values()
                detail_data = list(detail_data)
                return restful.ok(data=detail_data)
            except:
                detail_data = BudgetCodeForm.objects.filter(Id=detail_id).values()
                detail_data = list(detail_data)
                return restful.ok(data=detail_data)
        except:
            return restful.params_error(message='data fail ')

#生成报表文件的处理方式
def statement_bring_info(request):
    if request.method == "POST":
        statement_ids = request.POST.getlist('statement_ids[]')
        # sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where
        #                                             c.relname = '%s' and a.attrelid = c.oid and a.attnum>0''' % 'BudgetCodeForm'
        #表头信息
        tb1 =['部门','備註/新增或損耗','鏈接','開單狀況','預算編號','PMCS单号','申請日期','PMCS簽核日期','PIC'
            ,'設備名稱/治具類型','規格/型號/版本','類別','單價','申請數量','單位','總費用(RMB)','實際發生金額(KRMB)'
            ,'RMB/USD','客戶','機種','ProjectCode','申請原因/用途']
        tb2 = ['部门', '備註/新增或損耗', '鏈接', '開單狀況', '預算編號', '201单号', '申請日期', '201簽核日期', 'PIC'
            , '設備名稱/治具類型', '規格/型號/版本', '類別', '單價', '申請數量', '單位', '總費用(RMB)', '實際發生金額(KRMB)'
            , 'RMB/USD', '客戶', '機種', 'ProjectCode', '申請原因/用途']
        sql2 = 'select "Department","Remark","AttachmentPath","BillingType","BudgetCode","ExternalNumberType"' \
               ',"ExternalNumber","ApplyDate","ExternalNumberEffectiveDate","Pic","ProductName","Model"' \
               ',"PurchaseType","UnitPrice","Quantity","Unit","Currency","Customer","TypeOfMachine"' \
               ',"ProjectCode","ApplyReason" from "BudgetCodeForm" where'
        if len(statement_ids) == 1:
            statement_ids= statement_ids[0]
            sql3 = sql2 + '"Id"=' + statement_ids

        else:
            statement_ids = tuple(statement_ids)
            statement_ids =str(statement_ids)
            sql3 = sql2 + '"Id" in ' + statement_ids

        cur = connection.cursor()
        cur.execute(sql3)
        data = list(cur.fetchall())
        data1=[]
        data2=[]
        for i in data:
            i=list(i)
            if i[5] == '1':
                if i[3] == '0':
                    i[3] = "單獨開單"
                if i[3] == '1':
                    i[3] = "合併開單"
                i.remove('1')
                sum = int(i[12]) * int(i[13])
                i.insert(15,str(sum))
                if i[13] == "杂购":
                    i.insert(16,str(sum/1000))
                elif sum>600000:
                    i.insert(16,str(sum/3600))
                else:
                    i.insert(16,str(sum/1200))
                data1.append(i)
            if i[5] == '2':
                if i[3] == '0':
                    i[3] = "單獨開單"
                if i[3] == '1':
                    i[3] = "合併開單"
                i.remove('2')
                sum = int(i[12]) * int(i[13])
                i.insert(15, str(sum))
                if i[13] == "杂购":
                    i.insert(16, str(sum / 1000))
                elif sum > 600000:
                    i.insert(16, str(sum / 3600))
                else:
                    i.insert(16, str(sum / 1200))
                data2.append(i)

        data1.insert(0,tb1)
        data2.insert(0,tb2)

        #写入文件
        time_num = int(time.time())
        time_num = str(time_num)
        filename = 'download' + time_num + '.xlsx'
        wb = Workbook()
        index = 0
        sheet_name = "PMCS預算編號記錄"
        wb.create_sheet(sheet_name, index=index)
        sheet = wb[sheet_name]
        for row in data1:
            sheet.append(row)

        index = 1
        sheet_name = "201领用"
        wb.create_sheet(sheet_name, index=index)
        sheet = wb[sheet_name]
        for row2 in data2:
            sheet.append(row2)
        try:
            wb.save(os.path.join(settings.MEDIA_CHANGE_ROOT, filename))
            file_url = request.build_absolute_uri(settings.MEDIA_CHANGE_URL + filename)
            data = [file_url]
            return restful.ok(data=data)
        except:
            return restful.params_error(message=" please select one information ")


#生成报表并存入服务器的函数
def statement_excle(request,data,sheet_name,file_root,file_url,filename):
    # 写入文件
    # time_num = int(time.time())
    # time_num = str(time_num)
    # sheet_name = "预算表单1"
    # filename = 'download' + time_num + '.xlsx'

    wb = Workbook()
    index = 0
    wb.create_sheet(sheet_name, index=index)
    sheet = wb[sheet_name]
    for row in data:
        row = list(row)
        if len(row)==7:
            #正常
            if row[2] > row[4]:
                row.append("正常")
            #预警
            if row[2] <= row[4] <=row[3]:
                row.append("预警")
            #超标
            if row[3] < row[4]:
                row.append("超标")
        sheet.append(row)
    wb.save(os.path.join(file_root, filename))

    file_url = request.build_absolute_uri(file_url + filename)
    data = [file_url]
    return data


#设备NG率的监控类以及生成报表
class MonitorEquipment(View):

    def get(self,request):
        try:
            #这里默认是查询前几周的数据
            # end_time = datetime.datetime.now()
            # delta = datetime.timedelta(days=21)
            # start_time = end_time-delta
            # print(start_time.strftime('%Y-%m-%d %H:%M:%S'))
            #页面html显示需要的数据
            dict_data = {}
            data = PartItem.objects.order_by("Id").all()[0:10]
            limit_value = Configuration.objects.filter(Type="NG率设定").values("Max","Min","Id")
            limit_value = list(limit_value)
            data = list(data.values())
            dict_data['data']=data
            dict_data['limit_value']=limit_value
            #饼状图需要的数据
            sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where'
            try:
                parmeter_stands = Configuration.objects.get(Type="NG率设定")
                parmeter_stands_min = str(parmeter_stands.Min)
                parmeter_stands_max = str(parmeter_stands.Max)
                #正常
                sql1 = sql + '"NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
                #预警
                sql2 = sql + '"NGRate" >= \'' + parmeter_stands_min + '\'' \
                      + 'AND "NGRate" <= \'' + parmeter_stands_max + '\' group by "PartName"'
                #超标
                sql3 = sql + '"NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'


                cur = connection.cursor()
                cur.execute(sql1)
                normal = cur.fetchall()
                cur.execute(sql2)
                warning =cur.fetchall()
                cur.execute(sql3)
                danger = cur.fetchall()
                cur.close()
                dict_data['normal']=normal
                dict_data['warning']=warning
                dict_data['danger']=danger
            except:
                sql4 = sql+' 1=1'
                cur = connection.cursor()
                cur.execute(sql4)
                normal = cur.fetchall()
                dict_data['normal'] = normal
            return restful.ok(data=dict_data)
        except:
            return restful.params_error(message="")

    def post(self,request):
        NG_ids = request.POST.getlist('NG_ids[]')
        sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where
                                                    c.relname = '%s' and a.attrelid = c.oid and a.attnum>0''' % 'PartItem'
        sql2 = 'SELECT "SN","PartName","Configuration"."Min","Configuration"."Max","NGRate","ErrorCounts","UsedTimes" FROM "PartItem" INNER JOIN "Configuration" on "Configuration"."Id"=2 where '
        if len(NG_ids) == 1:
            NG_ids= NG_ids[0]
            sql3 = sql2 + '"PartItem"."Id"=' + NG_ids
        else:
            NG_ids = tuple(NG_ids)
            NG_ids =str(NG_ids)
            sql3 = sql2 + '"PartItem"."Id" in ' + NG_ids

        cur = connection.cursor()
        cur.execute(sql3)
        data = cur.fetchall()
        cur.close()

        cur2 = connection.cursor()
        cur2.execute(sql1)
        statement_data = cur2.fetchall()
        cur2.close()
        statement_data = [attr[0] for attr in statement_data]

        data1 = []
        data1.append(statement_data[statement_data.index("SN")])
        data1.append(statement_data[statement_data.index("PartName")])
        data1.append('Min')
        data1.append('Max')
        data1.append(statement_data[statement_data.index("NGRate")])
        data1.append(statement_data[statement_data.index("ErrorCounts")])
        data1.append(statement_data[statement_data.index("UsedTimes")])
        data1.append("status")

        data.insert(0,data1)
        try:
            sheet_name = "NG率监控表单"
            time_num = int(time.time())
            time_num = str(time_num)
            file_root =settings.MEDIA_MONITOR_ROOT
            file_url = settings.MEDIA_MONITOR_URL
            filename = 'download'+time_num+'.xlsx'
            data = statement_excle(request,data,sheet_name,file_root,file_url,filename)
            return restful.ok(data=data)

        except:
            return restful.params_error(message="download fail")
#设置预警区间
def setup_parameter(request):
    if request.method == "POST":
        min = request.POST['min']
        max = request.POST['max']
        mail_receiver = request.POST.getlist('mail_receiver[]')
        mail_receiver =list(mail_receiver)
        NG_monitor_type = "NG率设定"
        parmeter_id = Configuration.objects.filter(Type=NG_monitor_type)
        mail_receiver_count = ",".join(mail_receiver)
        try:
            if parmeter_id:
                Configuration.objects.filter(Type=NG_monitor_type).update(Max=max,Min=min,Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
            else:
                Configuration.objects.create(Type=NG_monitor_type, Max=max, Min=min, Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
        except:
            return restful.params_error(message="setup parameter fail")


#查询数据的信息 模糊查询数据
def monitor_query_info(request):
    if request.method == "POST":
        sn = request.POST['sn']
        part_name = request.POST['part_name']
        status = request.POST['status']
        NG_monitor_type = "NG率设定"
        sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
              ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
              'FROM "PartItem" WHERE 1=1'
        visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where 1=1 '
        sn = str(sn)
        try:
            parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
            limit_value =[parmeter_stands.Min,parmeter_stands.Max]
            parmeter_stands_min =str(parmeter_stands.Min)
            parmeter_stands_max =str(parmeter_stands.Max)
            dict_data = {}
            normal = []
            warning = []
            danger = []
            if status == "":
                if part_name !="":
                    sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                if sn !="":
                    sql = sql + 'AND "SN" = \'' + sn + '\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                # 正常
                visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
                # 预警
                visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                               + 'AND "NGRate" <= \'' + parmeter_stands_max + '\' group by "PartName"'
                # 超标
                visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
                cur = connection.cursor()
                cur.execute(visual_sql_normal)
                normal = cur.fetchall()

                cur = connection.cursor()
                cur.execute(visual_sql_waring)
                warning = cur.fetchall()

                cur = connection.cursor()
                cur.execute(visual_sql_danger)
                danger = cur.fetchall()

                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()

                dict_data['data'] = data
                dict_data['limit_value'] = limit_value
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger

                return restful.ok(data=dict_data)
            else:
                # global normal,warning,danger,dict_data
                if sn != "":
                    sql = sql+'AND "SN" = \''+ sn +'\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                if part_name != "":
                    sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                if status == "正常":
                    sql = sql+'AND "NGRate" < \''+ parmeter_stands_min +'\''
                    visual_sql_normal = visual_sql+'AND "NGRate" < \''+ parmeter_stands_min +'\' group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_normal)
                    normal = cur.fetchall()

                if status == "预警":
                    sql = sql + 'AND "NGRate" >= \''+ parmeter_stands_min +'\''+ 'AND "NGRate" <= \''+ parmeter_stands_max +'\''
                    visual_sql_waring = visual_sql+'AND "NGRate" >= \''+ parmeter_stands_min +'\''+ 'AND "NGRate" <= \''+ parmeter_stands_max +'\'group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_waring)
                    warning = cur.fetchall()

                if status == "超标":
                    sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                    visual_sql_danger = visual_sql+'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_danger)
                    danger = cur.fetchall()

                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
                dict_data['data']=data
                dict_data['limit_value']=limit_value
                return restful.ok(data=dict_data)
        except:
            return restful.params_error(message="query data is empty")

#html视图中点击事件
def visual_data(request):
    if request.method == "POST":
        try:
            status = request.POST['status']
            part_name = request.POST['part_name']
            NG_monitor_type = "NG率设定"
            dict_data = {}

            parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
            limit_value = [parmeter_stands.Min, parmeter_stands.Max]
            parmeter_stands_min = str(parmeter_stands.Min)
            parmeter_stands_max = str(parmeter_stands.Max)
            sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
                  ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
                  'FROM "PartItem" WHERE "PartName" = \'' + part_name + '\''

            if status == "#28a745":
                sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
            if status == "#ffc107":
                sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                      + 'AND "NGRate" <= \'' + parmeter_stands_max + '\''
            if status == "#dc3545":
                sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''

            cur = connection.cursor()
            cur.execute(sql)
            data = cur.fetchall()

            dict_data['data'] = data
            dict_data['limit_value'] = limit_value
            return restful.ok(data=dict_data)
        except:
            return restful.params_error(message="query data is empty")

#捞出所有NG率达到或超过预警区间的SN发邮件提醒给收件人 定时的功能在被使用在DBexcle app。views里面的函数crontab_test使用了
def check_NGRate():
    NG_monitor_type = "NG率设定"
    parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
    # limit_value = [parmeter_stands.Min, parmeter_stands.Max]
    parmeter_stands_min = str(parmeter_stands.Min)
    parmeter_stands_receiver = str(parmeter_stands.Reminders)
    sql_remenber = 'select "SN" from "PartItem" where "NGRate" >= \'' + parmeter_stands_min + '\''
    cur = connection.cursor()
    cur.execute(sql_remenber)
    data = cur.fetchall()
    data = str(data)
    if len(data) >0:
        email_1=[]
        receiver_list = parmeter_stands_receiver.split(',')
        for i in range(len(receiver_list)):
            email_1.append(receiver_list[i]+'@wistron.com')

        subject = "Notification: AEMSLite system remind information"
        content = """
<pre>
Dear """ + parmeter_stands_receiver + """ ,
    Follow this NG Rate monitor information, please see the below link address:
    
    NG Rate receive to stands or over data :""" + data + """ 
    Please click the link below to see the detail information:
    <a href="http://10.41.95.106:90/index">index AEMSLite</a>


    =====================================================================================================
    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
        mail.sendmail(email_1, content, subject)
    else:
        return restful.ok(message="partName is normal")


#保养类的函数
class maintain_equipment_info(View):
    def get(self,request):
        dict_data={}
        data = PartItem.objects.order_by("Id").all()[0:10]
        limit_value_1 = Configuration.objects.filter(Type="mt_count").values("Max", "Min", "Id")
        limit_value_2 = Configuration.objects.filter(Type="mt_date").values("Max", "Min", "Id")
        limit_value1 = list(limit_value_1)
        limit_value2 = list(limit_value_2)
        data = list(data.values())
        start_time = str(datetime.now()).split(' ')[0]
        start_time = datetime.strptime(start_time, "%Y-%m-%d")
        #计算保养次数和保养日期达不达标
        for i in range(0,len(data)):
            if data[i]['NextCheckDate'] == None and data[i]['NextCheckCount'] ==0:
                data[i]['stand_date'] = 0
                data[i]['stand_count'] = 0
            else:
                time_u = datetime.strptime(str(data[i]['NextCheckDate']).split(' ')[0], "%Y-%m-%d")
                days = time_u - start_time
                data[i]['stand_date'] = days.days
                data[i]['stand_count'] = data[i]['NextCheckCount'] - data[i]['UsedTimes']

        dict_data['data'] = data
        dict_data['limit_value1'] = limit_value1
        dict_data['limit_value2'] = limit_value2
        sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where 1=1 '

        # 饼状图需要的数据
        delta = timedelta(days=3)
        # end_time = datetime.datetime.now()
        # start_time = end_time-delta
        # end_time =str(end_time)
        # start_time =str(start_time)
        # visual_sql = sql+' AND "TrnDate" <=\'{0}\' AND "TrnDate" >=\'{1}\''.format(end_time,start_time)
        # cur = connection.cursor()
        # cur.execute(visual_sql)
        # visual_data = cur.fetchall()
        # dict_data['visual_data'] = visual_data

        parmeter_stands_1 = Configuration.objects.get(Type="mt_count")
        parmeter_stands_2 = Configuration.objects.get(Type="mt_date")
        parmeter_stands_1_max = str(int(parmeter_stands_1.Max))
        parmeter_stands_1_min = str(int(parmeter_stands_1.Max-5))
        parmeter_stands_2_max = str(int(parmeter_stands_2.Max))
        parmeter_stands_2_min = str(int(parmeter_stands_2.Max-float(delta.days)))

        # 正常
        sql1 = sql + 'AND "CheckCycleCount" < ' +parmeter_stands_1_min+' AND "CheckCycle" < ' +parmeter_stands_2_min+' GROUP BY "PartName"'
        # 预警
        sql2 = sql + 'AND "CheckCycleCount" >= ' +parmeter_stands_1_min+' AND "CheckCycle" >= ' +parmeter_stands_2_min
        sql2 = sql2+ 'AND "CheckCycleCount" <= ' +parmeter_stands_1_max+' AND "CheckCycle" <= ' +parmeter_stands_2_max+'  GROUP BY "PartName"'
        # 超标
        sql3 = sql + 'AND "CheckCycleCount" > '+parmeter_stands_1_max+' AND "CheckCycle" > '+parmeter_stands_2_max+'  GROUP BY "PartName"'
        cur = connection.cursor()
        cur.execute(sql1)
        normal = cur.fetchall()

        cur.execute(sql2)
        warning = cur.fetchall()

        cur.execute(sql3)
        danger = cur.fetchall()
        try:
            dict_data['normal'] = normal
            dict_data['warning'] = warning
            dict_data['danger'] = danger
            return restful.ok(data=dict_data)
        except:
            return restful.ok(data=dict_data)

    def post(self,request):
        maintain_count = request.POST['maintain_count']
        maintain_date = request.POST['maintain_date']
        maintain_receiver = request.POST.getlist('maintain_receiver[]')
        maintain_receiver = list(maintain_receiver)
        try:
            maintain_type_count = "mt_count"
            maintain_type_date = "mt_date"
            parameter_count = Configuration.objects.filter(Type=maintain_type_count)
            parameter_date = Configuration.objects.filter(Type=maintain_type_date)
            mail_receiver_count = ",".join(maintain_receiver)

            if parameter_count and parameter_date:
                Configuration.objects.filter(Type=maintain_type_count).update(Max=maintain_count,Min=0,Reminders=mail_receiver_count)
                Configuration.objects.filter(Type=parameter_date).update(Max=maintain_date,Min=0,Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
            else:
                Configuration.objects.create(Type=maintain_type_count, Max=maintain_count, Min=0, Reminders=mail_receiver_count)
                Configuration.objects.create(Type=maintain_type_date, Max=maintain_date, Min=0, Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter create success")
        except:
            return restful.params_error(message="setup information error")

#单独的SN的保养更改
def maintain_setup_info(request):
    if request.method == "POST":
        main_count = request.POST['main_count']
        main_cycle = request.POST['main_cycle']
        main_date = request.POST['main_date']
        main_sn = request.POST['main_sn']
        try:
            PartItem.objects.filter(SN=main_sn).update(CheckCycleCount=main_count,CheckCycle=main_cycle,NextCheckDate=main_date)
            return restful.ok(message="maintain modify success")
        except:
            return restful.params_error(message="maintain modify fail")

#by PN的批量更改
def maintain_setup_by_pn(request):
    if request.method == "POST":
        main_pn = request.POST['main_partname']
        main_count = request.POST['main_count']
        main_day = request.POST['main_day']
        main_date = request.POST['main_date']
        try:
            result = PartItem.objects.filter(PN=main_pn)
            if len(restful) >0:
                result.update(CheckCycleCount=main_count,CheckCycle=main_day,NextCheckDate=main_date)
                return restful.ok(message="maintain modify success")
            return restful.params_error(message="PN query is null")
        except:
            return restful.params_error(message="PN query is null")
















#统计分析的数据的 拉出一周的数据， 这里先拉出来前面10条的数据
class analysis_equipment_info(View):
    def get(self,request):
        try:
            visua_data = {}
            #柱状图需要的数据
            sql1 = 'SELECT "ErrorCode", COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' GROUP BY "ErrorCode"'
            cur = connection.cursor()
            cur.execute(sql1)
            visua_data['errorcode'] = cur.fetchall()

            sql2 = 'SELECT "PartName", COUNT("SN") FROM "PartItemResult" GROUP BY "PartName"'
            cur = connection.cursor()
            cur.execute(sql2)
            visua_data['Partname'] = cur.fetchall()
            try:
                sql3 = 'SELECT COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' and "UsedTimes">0'
                cur = connection.cursor()
                cur.execute(sql3)
                visua_data['user'] = cur.fetchall()
            except:
                sql3 = 'SELECT COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' and "UsedTimes">0'
                cur = connection.cursor()
                cur.execute(sql3)
                visua_data['user'] = cur.fetchall()
            sql4 = 'SELECT COUNT("SN"),"PartName" FROM (select distinct "SN","PartName","Result" from "PartItemResult") as foo where "Result"= \'FAIL\' GROUP BY "PartName"'
            cur = connection.cursor()
            cur.execute(sql4)
            visua_data['filterSN'] = cur.fetchall()

            return restful.ok(data=visua_data)
        except:
            return restful.params_error(message='data error')

    def post(self,request):
        pass


#数据显示部分
def analysis_data(request):
    if request.method == "GET":
        #数据显示部分的data
        try:
            data = PartItemResult.objects.order_by().all().values()
            data =list(data)
            return restful.ok(data=data)
        except:
            return restful.params_error(message="data got fail")

#设置区间的获取的数据
def analysis_setup_data(request):
    if request.method == "GET":
        try:
            limit_data = Configuration.objects.filter(Type="sa_block").values("Id","Min","Max")
            limit_data = list(limit_data)
            return restful.ok(data=limit_data)
        except:
            return restful.params_error(message="data error")

#设置区间提交的数据摄入表里面
def analysis_setup_value(request):
    if request.method == "POST":
        try:
            return restful.ok(message="setup success")
        except:
            return restful.params_error(message="data error")